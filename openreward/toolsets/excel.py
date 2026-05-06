from __future__ import annotations

import json
from typing import Literal, Optional
from pydantic import BaseModel, Field

from openreward.environments.environment import tool
from openreward.environments.toolset import Toolset
from openreward.environments.types import TextBlock, ToolOutput


# ===== Pydantic 参数模型 =====

class CreateSpreadsheetParams(BaseModel):
    file_path: str = Field(..., description="新电子表格的创建路径")
    sheet_name: str = Field("Sheet1", description="初始工作表名称")


class DeleteSpreadsheetParams(BaseModel):
    file_path: str = Field(..., description="要删除的 Excel 电子表格路径")


class ListTabsParams(BaseModel):
    file_path: str = Field(..., description="Excel 电子表格路径")


class AddTabParams(BaseModel):
    file_path: str = Field(..., description="Excel 电子表格路径")
    sheet_name: str = Field(..., description="新工作表名称")
    position: int | None = Field(None, description="位置索引（从 0 开始），None 表示追加到末尾")


class DeleteTabParams(BaseModel):
    file_path: str = Field(..., description="Excel 电子表格路径")
    sheet_name: str = Field(..., description="要删除的工作表名称")


class ReadTabParams(BaseModel):
    file_path: str = Field(..., description="Excel 电子表格路径")
    sheet_name: str = Field(..., description="要读取的工作表名称")
    max_rows: int | None = Field(None, description="最多读取行数")
    max_cols: int | None = Field(None, description="最多读取列数")


class ReadCsvParams(BaseModel):
    csv_path: str = Field(..., description="CSV 文件路径")
    max_rows: int | None = Field(None, description="最多读取行数")


class EditSpreadsheetParams(BaseModel):
    file_path: str = Field(..., description="Excel 电子表格路径")
    sheet_name: str = Field(..., description="工作表名称")
    cell_reference: str = Field(..., description="单元格引用（例如，'A1'、'B5'）")
    value: str | int | float = Field(..., description="要写入单元格的值")


class AddContentTextParams(BaseModel):
    file_path: str = Field(..., description="Excel 电子表格路径")
    sheet_name: str = Field(..., description="工作表名称")
    start_cell: str = Field(..., description="起始单元格引用（例如，'A1'）")
    data: list[list] = Field(..., description="要写入的二维值数组 [行][列]")


class DeleteContentCellParams(BaseModel):
    file_path: str = Field(..., description="Excel 电子表格路径")
    sheet_name: str = Field(..., description="工作表名称")
    cell_reference: str = Field(..., description="要清空的单元格引用（例如，'A1'）")


class CreateChartParams(BaseModel):
    file_path: str = Field(..., description="Excel 电子表格路径")
    sheet_name: str = Field(..., description="工作表名称")
    chart_type: Literal["bar", "column", "line", "pie", "scatter", "area"] = Field(..., description="图表类型")
    data_range: str = Field(..., description="图表数据范围（例如，'A1:B10'）")
    chart_position: str = Field(..., description="图表左上角的单元格引用（例如，'D1'）")
    title: str | None = Field(None, description="图表标题")
    x_axis_title: str | None = Field(None, description="X 轴标签")
    y_axis_title: str | None = Field(None, description="Y 轴标签")


# ===== Excel 工具集类 =====

class ExcelToolset(Toolset):
    """通过 openpyxl 在沙箱中执行，提供 Excel 电子表格操作工具的工具集"""

    @tool
    async def excel_create_spreadsheet(self, params: CreateSpreadsheetParams) -> ToolOutput:
        """创建包含初始工作表的新 Excel 工作簿"""
        sheet_name_escaped = params.sheet_name.replace('\\', '\\\\').replace('"', '\\"')

        script = f'''
import json
from openpyxl import Workbook

try:
    wb = Workbook()
    ws = wb.active
    ws.title = "{sheet_name_escaped}"

    wb.save("{params.file_path}")

    result = {{
        "success": True,
        "file_path": "{params.file_path}",
        "sheet_name": "{sheet_name_escaped}",
    }}
    print(json.dumps(result))
except Exception as e:
    print(json.dumps({{"success": False, "error": str(e)}}))
'''

        cmd = f"python3 << 'PYSCRIPT_EOF'\n{script}\nPYSCRIPT_EOF"
        output, exit_code = await self.sandbox.run(cmd, max_bytes=1_000_000)
        try:
            result = json.loads(output)

            if not result.get("success"):
                return ToolOutput(
                    blocks=[TextBlock(text=f"❌ Error: {result.get('error')}")],
                    metadata={"error": result.get("error")},
                    reward=0.0,
                    finished=False,
                )

            return ToolOutput(
                blocks=[TextBlock(text=f"✅ Spreadsheet created successfully at {params.file_path}")],
                metadata=result,
                reward=0.0,
                finished=False,
            )
        except json.JSONDecodeError:
            return ToolOutput(
                blocks=[TextBlock(text=f"❌ Error parsing output: {output}")],
                metadata={"error": "JSON decode failed", "output": output},
                reward=0.0,
                finished=False,
            )

    @tool
    async def excel_delete_spreadsheet(self, params: DeleteSpreadsheetParams) -> ToolOutput:
        """从沙箱中删除 Excel 电子表格文件"""
        output, exit_code = await self.sandbox.run(f"rm {params.file_path}", max_bytes=1_000_000)

        if exit_code == 0:
            return ToolOutput(
                blocks=[TextBlock(text=f"✅ Spreadsheet deleted successfully: {params.file_path}")],
                metadata={"file_path": params.file_path, "deleted": True},
                reward=0.0,
                finished=False,
            )
        else:
            return ToolOutput(
                blocks=[TextBlock(text=f"❌ Failed to delete: {output}")],
                metadata={"file_path": params.file_path, "deleted": False, "error": output},
                reward=0.0,
                finished=False,
            )

    @tool
    async def excel_list_tabs_in_spreadsheet(self, params: ListTabsParams) -> ToolOutput:
        """列出工作簿中的所有工作表/标签名称"""
        script = f'''
import json
from openpyxl import load_workbook

try:
    wb = load_workbook("{params.file_path}")
    sheet_names = wb.sheetnames

    result = {{
        "success": True,
        "file_path": "{params.file_path}",
        "sheet_count": len(sheet_names),
        "sheet_names": sheet_names,
    }}
    print(json.dumps(result))
except Exception as e:
    print(json.dumps({{"success": False, "error": str(e)}}))
'''

        cmd = f"python3 << 'PYSCRIPT_EOF'\n{script}\nPYSCRIPT_EOF"
        output, exit_code = await self.sandbox.run(cmd, max_bytes=1_000_000)
        try:
            result = json.loads(output)

            if not result.get("success"):
                return ToolOutput(
                    blocks=[TextBlock(text=f"❌ Error: {result.get('error')}")],
                    metadata={"error": result.get("error")},
                    reward=0.0,
                    finished=False,
                )

            # 构建摘要
            summary = f"Spreadsheet: {params.file_path}\n"
            summary += f"Total sheets: {result['sheet_count']}\n\n"
            summary += "Sheets:\n"
            for idx, name in enumerate(result['sheet_names']):
                summary += f"  [{idx}] {name}\n"

            return ToolOutput(
                blocks=[TextBlock(text=summary)],
                metadata=result,
                reward=0.0,
                finished=False,
            )
        except json.JSONDecodeError:
            return ToolOutput(
                blocks=[TextBlock(text=f"❌ Error parsing output: {output}")],
                metadata={"error": "JSON decode failed", "output": output},
                reward=0.0,
                finished=False,
            )

    @tool
    async def excel_add_tab(self, params: AddTabParams) -> ToolOutput:
        """向现有工作簿添加新工作表"""
        sheet_name_escaped = params.sheet_name.replace('\\', '\\\\').replace('"', '\\"')
        position_code = f"{params.position}" if params.position is not None else "None"

        script = f'''
import json
from openpyxl import load_workbook

try:
    wb = load_workbook("{params.file_path}")

    # 在指定位置或末尾创建新工作表
    position = {position_code}
    if position is not None:
        wb.create_sheet(title="{sheet_name_escaped}", index=position)
    else:
        wb.create_sheet(title="{sheet_name_escaped}")

    wb.save("{params.file_path}")

    result = {{
        "success": True,
        "file_path": "{params.file_path}",
        "sheet_name": "{sheet_name_escaped}",
        "sheet_count": len(wb.sheetnames),
    }}
    print(json.dumps(result))
except Exception as e:
    print(json.dumps({{"success": False, "error": str(e)}}))
'''

        cmd = f"python3 << 'PYSCRIPT_EOF'\n{script}\nPYSCRIPT_EOF"
        output, exit_code = await self.sandbox.run(cmd, max_bytes=1_000_000)
        try:
            result = json.loads(output)

            if not result.get("success"):
                return ToolOutput(
                    blocks=[TextBlock(text=f"❌ Error: {result.get('error')}")],
                    metadata={"error": result.get("error")},
                    reward=0.0,
                    finished=False,
                )

            return ToolOutput(
                blocks=[TextBlock(text=f"✅ Sheet '{params.sheet_name}' added successfully (total: {result['sheet_count']} sheets)")],
                metadata=result,
                reward=0.0,
                finished=False,
            )
        except json.JSONDecodeError:
            return ToolOutput(
                blocks=[TextBlock(text=f"❌ Error parsing output: {output}")],
                metadata={"error": "JSON decode failed", "output": output},
                reward=0.0,
                finished=False,
            )

    @tool
    async def excel_delete_tab(self, params: DeleteTabParams) -> ToolOutput:
        """从工作簿中移除工作表"""
        sheet_name_escaped = params.sheet_name.replace('\\', '\\\\').replace('"', '\\"')

        script = f'''
import json
from openpyxl import load_workbook

try:
    wb = load_workbook("{params.file_path}")

    if "{sheet_name_escaped}" not in wb.sheetnames:
        print(json.dumps({{
            "success": False,
            "error": f"Sheet '{{sheet_name_escaped}}' not found. Available sheets: {{wb.sheetnames}}"
        }}))
    else:
        ws = wb["{sheet_name_escaped}"]
        wb.remove(ws)
        wb.save("{params.file_path}")

        result = {{
            "success": True,
            "file_path": "{params.file_path}",
            "deleted_sheet": "{sheet_name_escaped}",
            "remaining_sheets": len(wb.sheetnames),
        }}
        print(json.dumps(result))
except Exception as e:
    print(json.dumps({{"success": False, "error": str(e)}}))
'''

        cmd = f"python3 << 'PYSCRIPT_EOF'\n{script}\nPYSCRIPT_EOF"
        output, exit_code = await self.sandbox.run(cmd, max_bytes=1_000_000)
        try:
            result = json.loads(output)

            if not result.get("success"):
                return ToolOutput(
                    blocks=[TextBlock(text=f"❌ Error: {result.get('error')}")],
                    metadata={"error": result.get("error")},
                    reward=0.0,
                    finished=False,
                )

            return ToolOutput(
                blocks=[TextBlock(text=f"✅ Sheet '{params.sheet_name}' deleted successfully ({result['remaining_sheets']} sheets remaining)")],
                metadata=result,
                reward=0.0,
                finished=False,
            )
        except json.JSONDecodeError:
            return ToolOutput(
                blocks=[TextBlock(text=f"❌ Error parsing output: {output}")],
                metadata={"error": "JSON decode failed", "output": output},
                reward=0.0,
                finished=False,
            )

    @tool
    async def excel_read_tab(self, params: ReadTabParams) -> ToolOutput:
        """从特定工作表读取数据"""
        sheet_name_escaped = params.sheet_name.replace('\\', '\\\\').replace('"', '\\"')
        max_rows = params.max_rows if params.max_rows else 999999
        max_cols = params.max_cols if params.max_cols else 999999

        script = f'''
import json
import datetime
from openpyxl import load_workbook

def default_serializer(obj):
    if isinstance(obj, (datetime.datetime, datetime.date, datetime.time)):
        return obj.isoformat()
    if isinstance(obj, datetime.timedelta):
        return str(obj)
    raise TypeError(f"Object of type {{type(obj).__name__}} is not JSON serializable")

try:
    wb = load_workbook("{params.file_path}")

    if "{sheet_name_escaped}" not in wb.sheetnames:
        print(json.dumps({{
            "success": False,
            "error": f"Sheet '{{sheet_name_escaped}}' not found. Available sheets: {{wb.sheetnames}}"
        }}))
    else:
        ws = wb["{sheet_name_escaped}"]

        # 从工作表读取数据
        data = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx >= {max_rows}:
                break
            row_data = list(row[:min(len(row), {max_cols})])
            data.append(row_data)

        result = {{
            "success": True,
            "file_path": "{params.file_path}",
            "sheet_name": "{sheet_name_escaped}",
            "rows": len(data),
            "cols": len(data[0]) if data else 0,
            "data": data,
        }}
        print(json.dumps(result, default=default_serializer))
except Exception as e:
    print(json.dumps({{"success": False, "error": str(e)}}))
'''

        cmd = f"python3 << 'PYSCRIPT_EOF'\n{script}\nPYSCRIPT_EOF"
        output, exit_code = await self.sandbox.run(cmd, max_bytes=1_000_000)
        try:
            result = json.loads(output)

            if not result.get("success"):
                return ToolOutput(
                    blocks=[TextBlock(text=f"❌ Error: {result.get('error')}")],
                    metadata={"error": result.get("error")},
                    reward=0.0,
                    finished=False,
                )

            # 构建带预览的摘要
            summary = f"Read sheet '{params.sheet_name}' from {params.file_path}\n"
            summary += f"Dimensions: {result['rows']} rows × {result['cols']} columns\n\n"
            summary += "Data:\n"
            for row_idx, row in enumerate(result['data']):
                row_str = ", ".join([str(cell) if cell is not None else "(empty)" for cell in row])
                summary += f"  Row {row_idx + 1}: {row_str}\n"

            return ToolOutput(
                blocks=[TextBlock(text=summary)],
                metadata=result,
                reward=0.0,
                finished=False,
            )
        except json.JSONDecodeError:
            return ToolOutput(
                blocks=[TextBlock(text=f"❌ Error parsing output: {output}")],
                metadata={"error": "JSON decode failed", "output": output},
                reward=0.0,
                finished=False,
            )

    @tool
    async def excel_read_csv(self, params: ReadCsvParams) -> ToolOutput:
        """读取 CSV 文件并转换为 Excel 格式数据结构"""
        max_rows = params.max_rows if params.max_rows else 999999

        script = f'''
import json
import csv

try:
    data = []
    with open("{params.csv_path}", 'r') as f:
        reader = csv.reader(f)
        for row_idx, row in enumerate(reader):
            if row_idx >= {max_rows}:
                break
            data.append(row)

    result = {{
        "success": True,
        "csv_path": "{params.csv_path}",
        "rows": len(data),
        "cols": len(data[0]) if data else 0,
        "data": data,
    }}
    print(json.dumps(result))
except Exception as e:
    print(json.dumps({{"success": False, "error": str(e)}}))
'''

        cmd = f"python3 << 'PYSCRIPT_EOF'\n{script}\nPYSCRIPT_EOF"
        output, exit_code = await self.sandbox.run(cmd, max_bytes=1_000_000)
        try:
            result = json.loads(output)

            if not result.get("success"):
                return ToolOutput(
                    blocks=[TextBlock(text=f"❌ Error: {result.get('error')}")],
                    metadata={"error": result.get("error")},
                    reward=0.0,
                    finished=False,
                )

            # 构建带预览的摘要
            summary = f"Read CSV file: {params.csv_path}\n"
            summary += f"Dimensions: {result['rows']} rows × {result['cols']} columns\n\n"
            summary += "Data:\n"
            for row_idx, row in enumerate(result['data']):
                row_str = ", ".join(row)
                summary += f"  Row {row_idx + 1}: {row_str}\n"

            return ToolOutput(
                blocks=[TextBlock(text=summary)],
                metadata=result,
                reward=0.0,
                finished=False,
            )
        except json.JSONDecodeError:
            return ToolOutput(
                blocks=[TextBlock(text=f"❌ Error parsing output: {output}")],
                metadata={"error": "JSON decode failed", "output": output},
                reward=0.0,
                finished=False,
            )

    @tool
    async def excel_edit_spreadsheet(self, params: EditSpreadsheetParams) -> ToolOutput:
        """修改工作表中现有单元格的值"""
        sheet_name_escaped = params.sheet_name.replace('\\', '\\\\').replace('"', '\\"')

        # 处理不同类型的值
        if isinstance(params.value, str):
            value_str = params.value.replace('\\', '\\\\').replace('"', '\\"')
            value_code = f'"{value_str}"'
        else:
            value_code = str(params.value)

        script = f'''
import json
from openpyxl import load_workbook

try:
    wb = load_workbook("{params.file_path}")

    if "{sheet_name_escaped}" not in wb.sheetnames:
        print(json.dumps({{
            "success": False,
            "error": f"Sheet '{{sheet_name_escaped}}' not found. Available sheets: {{wb.sheetnames}}"
        }}))
    else:
        ws = wb["{sheet_name_escaped}"]
        old_value = ws["{params.cell_reference}"].value

        ws["{params.cell_reference}"] = {value_code}

        wb.save("{params.file_path}")

        result = {{
            "success": True,
            "file_path": "{params.file_path}",
            "sheet_name": "{sheet_name_escaped}",
            "cell_reference": "{params.cell_reference}",
            "old_value": str(old_value) if old_value is not None else None,
            "new_value": str(ws["{params.cell_reference}"].value),
        }}
        print(json.dumps(result))
except Exception as e:
    print(json.dumps({{"success": False, "error": str(e)}}))
'''

        cmd = f"python3 << 'PYSCRIPT_EOF'\n{script}\nPYSCRIPT_EOF"
        output, exit_code = await self.sandbox.run(cmd, max_bytes=1_000_000)
        try:
            result = json.loads(output)

            if not result.get("success"):
                return ToolOutput(
                    blocks=[TextBlock(text=f"❌ Error: {result.get('error')}")],
                    metadata={"error": result.get("error")},
                    reward=0.0,
                    finished=False,
                )

            return ToolOutput(
                blocks=[TextBlock(text=f"✅ Cell {params.cell_reference} updated in sheet '{params.sheet_name}'\nNew value: {result['new_value']}")],
                metadata=result,
                reward=0.0,
                finished=False,
            )
        except json.JSONDecodeError:
            return ToolOutput(
                blocks=[TextBlock(text=f"❌ Error parsing output: {output}")],
                metadata={"error": "JSON decode failed", "output": output},
                reward=0.0,
                finished=False,
            )

    @tool
    async def excel_add_content_text(self, params: AddContentTextParams) -> ToolOutput:
        """将数据写入单元格范围（批量操作）"""
        sheet_name_escaped = params.sheet_name.replace('\\', '\\\\').replace('"', '\\"')
        data_json = json.dumps(params.data)

        script = f'''
import json
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

try:
    wb = load_workbook("{params.file_path}")

    if "{sheet_name_escaped}" not in wb.sheetnames:
        print(json.dumps({{
            "success": False,
            "error": f"Sheet '{{sheet_name_escaped}}' not found. Available sheets: {{wb.sheetnames}}"
        }}))
    else:
        ws = wb["{sheet_name_escaped}"]
        data = {data_json}

        # 解析起始单元格（例如，"A1" -> 行=1, 列=1）
        cell_ref = "{params.start_cell}"
        col_letter = ""
        row_num = ""
        for char in cell_ref:
            if char.isalpha():
                col_letter += char
            else:
                row_num += char

        start_row = int(row_num)
        start_col = column_index_from_string(col_letter)

        # 写入数据
        for row_idx, row_data in enumerate(data):
            for col_idx, value in enumerate(row_data):
                ws.cell(row=start_row + row_idx, column=start_col + col_idx, value=value)

        wb.save("{params.file_path}")

        result = {{
            "success": True,
            "file_path": "{params.file_path}",
            "sheet_name": "{sheet_name_escaped}",
            "start_cell": "{params.start_cell}",
            "rows_written": len(data),
            "cols_written": len(data[0]) if data else 0,
        }}
        print(json.dumps(result))
except Exception as e:
    print(json.dumps({{"success": False, "error": str(e)}}))
'''

        cmd = f"python3 << 'PYSCRIPT_EOF'\n{script}\nPYSCRIPT_EOF"
        output, exit_code = await self.sandbox.run(cmd, max_bytes=1_000_000)
        try:
            result = json.loads(output)

            if not result.get("success"):
                return ToolOutput(
                    blocks=[TextBlock(text=f"❌ Error: {result.get('error')}")],
                    metadata={"error": result.get("error")},
                    reward=0.0,
                    finished=False,
                )

            return ToolOutput(
                blocks=[TextBlock(text=f"✅ Data written to sheet '{params.sheet_name}' starting at {params.start_cell}\nDimensions: {result['rows_written']} rows × {result['cols_written']} columns")],
                metadata=result,
                reward=0.0,
                finished=False,
            )
        except json.JSONDecodeError:
            return ToolOutput(
                blocks=[TextBlock(text=f"❌ Error parsing output: {output}")],
                metadata={"error": "JSON decode failed", "output": output},
                reward=0.0,
                finished=False,
            )

    @tool
    async def excel_delete_content_cell(self, params: DeleteContentCellParams) -> ToolOutput:
        """清空单元格内容（设为 None）"""
        sheet_name_escaped = params.sheet_name.replace('\\', '\\\\').replace('"', '\\"')

        script = f'''
import json
from openpyxl import load_workbook

try:
    wb = load_workbook("{params.file_path}")

    if "{sheet_name_escaped}" not in wb.sheetnames:
        print(json.dumps({{
            "success": False,
            "error": f"Sheet '{{sheet_name_escaped}}' not found. Available sheets: {{wb.sheetnames}}"
        }}))
    else:
        ws = wb["{sheet_name_escaped}"]
        old_value = ws["{params.cell_reference}"].value

        ws["{params.cell_reference}"] = None

        wb.save("{params.file_path}")

        result = {{
            "success": True,
            "file_path": "{params.file_path}",
            "sheet_name": "{sheet_name_escaped}",
            "cell_reference": "{params.cell_reference}",
            "deleted_value": str(old_value) if old_value is not None else "(empty)",
        }}
        print(json.dumps(result))
except Exception as e:
    print(json.dumps({{"success": False, "error": str(e)}}))
'''

        cmd = f"python3 << 'PYSCRIPT_EOF'\n{script}\nPYSCRIPT_EOF"
        output, exit_code = await self.sandbox.run(cmd, max_bytes=1_000_000)
        try:
            result = json.loads(output)

            if not result.get("success"):
                return ToolOutput(
                    blocks=[TextBlock(text=f"❌ Error: {result.get('error')}")],
                    metadata={"error": result.get("error")},
                    reward=0.0,
                    finished=False,
                )

            return ToolOutput(
                blocks=[TextBlock(text=f"✅ Cell {params.cell_reference} cleared in sheet '{params.sheet_name}'\nDeleted value: {result['deleted_value']}")],
                metadata=result,
                reward=0.0,
                finished=False,
            )
        except json.JSONDecodeError:
            return ToolOutput(
                blocks=[TextBlock(text=f"❌ Error parsing output: {output}")],
                metadata={"error": "JSON decode failed", "output": output},
                reward=0.0,
                finished=False,
            )

    @tool
    async def excel_create_chart(self, params: CreateChartParams) -> ToolOutput:
        """在工作表中创建图表"""
        sheet_name_escaped = params.sheet_name.replace('\\', '\\\\').replace('"', '\\"')
        title_escaped = params.title.replace('\\', '\\\\').replace('"', '\\"') if params.title else ""
        x_axis_escaped = params.x_axis_title.replace('\\', '\\\\').replace('"', '\\"') if params.x_axis_title else ""
        y_axis_escaped = params.y_axis_title.replace('\\', '\\\\').replace('"', '\\"') if params.y_axis_title else ""

        # 将图表类型映射到 openpyxl 图表类
        chart_type_map = {
            "bar": "BarChart",
            "column": "BarChart",  # 不同方向的 BarChart
            "line": "LineChart",
            "pie": "PieChart",
            "scatter": "ScatterChart",
            "area": "AreaChart",
        }
        chart_class = chart_type_map.get(params.chart_type, "BarChart")

        # 构造包含工作表名称的完整引用字符串
        # 如果工作表名称包含空格或特殊字符，则引用
        if ' ' in params.sheet_name or any(c in params.sheet_name for c in ['!', "'", '"']):
            full_data_ref = f"'{params.sheet_name}'!{params.data_range}"
        else:
            full_data_ref = f"{params.sheet_name}!{params.data_range}"

        script = f'''
import json
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, AreaChart, Reference

try:
    wb = load_workbook("{params.file_path}")

    if "{sheet_name_escaped}" not in wb.sheetnames:
        print(json.dumps({{
            "success": False,
            "error": f"Sheet '{{sheet_name_escaped}}' not found. Available sheets: {{wb.sheetnames}}"
        }}))
    else:
        ws = wb["{sheet_name_escaped}"]

        # 创建图表
        chart = {chart_class}()

        # 使用包含工作表名称的完整引用字符串
        data = Reference(ws, range_string="{full_data_ref}")

        chart.add_data(data, titles_from_data=True)

        # 设置图表属性
        if "{title_escaped}":
            chart.title = "{title_escaped}"
        if "{x_axis_escaped}":
            chart.x_axis.title = "{x_axis_escaped}"
        if "{y_axis_escaped}":
            chart.y_axis.title = "{y_axis_escaped}"

        # 对于柱形图，设置类型为 column
        if "{params.chart_type}" == "column":
            chart.type = "col"

        # 将图表添加到工作表
        ws.add_chart(chart, "{params.chart_position}")

        wb.save("{params.file_path}")

        result = {{
            "success": True,
            "file_path": "{params.file_path}",
            "sheet_name": "{sheet_name_escaped}",
            "chart_type": "{params.chart_type}",
            "data_range": "{params.data_range}",
            "chart_position": "{params.chart_position}",
            "title": "{title_escaped}" if "{title_escaped}" else None,
        }}
        print(json.dumps(result))
except Exception as e:
    print(json.dumps({{"success": False, "error": str(e)}}))
'''

        cmd = f"python3 << 'PYSCRIPT_EOF'\n{script}\nPYSCRIPT_EOF"
        output, exit_code = await self.sandbox.run(cmd, max_bytes=1_000_000)
        try:
            result = json.loads(output)

            if not result.get("success"):
                return ToolOutput(
                    blocks=[TextBlock(text=f"❌ Error: {result.get('error')}")],
                    metadata={"error": result.get("error")},
                    reward=0.0,
                    finished=False,
                )

            chart_desc = f"{params.chart_type.capitalize()} chart"
            if params.title:
                chart_desc += f" '{params.title}'"

            return ToolOutput(
                blocks=[TextBlock(text=f"✅ {chart_desc} created in sheet '{params.sheet_name}' at {params.chart_position}")],
                metadata=result,
                reward=0.0,
                finished=False,
            )
        except json.JSONDecodeError:
            return ToolOutput(
                blocks=[TextBlock(text=f"❌ Error parsing output: {output}")],
                metadata={"error": "JSON decode failed", "output": output},
                reward=0.0,
                finished=False,
            )
